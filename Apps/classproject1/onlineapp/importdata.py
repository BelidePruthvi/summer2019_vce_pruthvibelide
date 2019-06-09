import os
import django
import openpyxl as xl
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject1.settings')
django.setup()
from onlineapp.models import *
print(Student.objects.all())


# def storedata():
#     workbook=xl.load_workbook(filename="students.xlsx")
#     worksheet=workbook['Colleges']
#     flag=0
#     for row in worksheet.rows:
#         if flag==0:
#             flag=1
#             continue
#         college=[cell.value for cell in row]
#         c=College(name=college[0],location=college[2],acronym=college[1],contact=college[3]);
#         c.save()
#
# def storestudent():
#     workbook = xl.load_workbook(filename="students.xlsx")
#     worksheet = workbook['Current']
#     flag=0
#     for row in worksheet.rows:
#         if flag==0:
#             flag=1
#             continue
#         student = [cell.value for cell in row]
#         c = Student(name=student[0],email=student[2],db_folder=student[3],college=College.objects.get(acronym=student[1]),dropped_out=True)
#         c.save()
#     workbook = xl.load_workbook(filename="students.xlsx")
#     worksheet = workbook['Deletions']
#     flag = 0
#     for row in worksheet.rows:
#         if flag == 0:
#             flag = 1
#             continue
#         student = [cell.value for cell in row]
#         clg = College.objects.get(acronym=student[1])
#         print('helo', clg)
#         c = Student(name=student[0], email=student[2], db_folder=student[3],
#                     college=College.objects.get(acronym=student[1]),dropped_out=False)
#         c.save()
#
# def mockResults():
#     workbook = xl.load_workbook(filename="out.xlsx")
#     worksheet = workbook['Test resul']
#     flag = 0
#     for row in worksheet.rows:
#         if flag == 0:
#             flag = 1
#             continue
#         values = [cell.value for cell in row]
#         temp = values[0].split('_')[-2]
#         c =MockTest1(problem1=values[1],problem2=values[2],problem3=values[3],problem4=values[4],total=values[5],student=Student.objects.get(db_folder = temp))
#         c.save()
#
# #storedata()
# #storestudent()
# mockResults()